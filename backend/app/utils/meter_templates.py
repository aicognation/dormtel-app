"""Generate pre-populated meter reading Excel templates (adhoc / daily / monthly).

These builders produce workbooks in the exact hardened formats that the existing
upload endpoints accept (see FIX-016/017):

- ``build_row_template``  -> the "one row per reading" format (sheet ``Meter Readings``)
  consumed by the Standard/Ad-hoc upload. Used for both Ad-hoc (Reading Date blank)
  and Daily (Reading Date pre-filled) templates.
- ``build_monthly_template`` -> the "one column per day" grid (sheet ``METER READINGS``)
  consumed by the Daily Sheet / Monthly Grid upload.

All builders accept the ``(Resident, Bed, Room)`` tuples returned by the billing
router's resident query so population rules stay server-side.
"""
import io
import calendar
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ── Brand styling ──
NAVY = "1F3A5F"
GOLD = "F2B705"
FILL_IN = "FFF8E1"  # light yellow = "type here"
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
GOLD_FILL = PatternFill("solid", fgColor=GOLD)
FILLIN_FILL = PatternFill("solid", fgColor=FILL_IN)
BASE_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
_THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# Canonical columns for the row-per-reading format (must match the Standard validator).
STD_HEADERS = ["Branch Code", "Building", "Room Number", "Bed", "Resident Name",
               "Reading Date (YYYY-MM-DD)", "Electric Reading (kWh)", "Water Reading (m³)"]


def _bed_letter(bed):
    if bed is not None and getattr(bed, "bed_code", None):
        return bed.bed_code[-1].upper()
    return ""


def _rate(resident, bed):
    val = getattr(bed, "rate_per_bed", None)
    if val is None:
        val = getattr(resident, "monthly_rate", None)
    try:
        return float(val) if val is not None else None
    except (ValueError, TypeError):
        return None


def _rows(resident_rows):
    """Normalize (Resident, Bed, Room) tuples into plain dicts for the builders."""
    out = []
    for resident, bed, room in resident_rows:
        out.append({
            "room": room.room_number.strip() if room is not None and room.room_number else "",
            "bed": _bed_letter(bed),
            "name": (resident.full_name or "").strip(),
            "rate": _rate(resident, bed),
            "move_in": resident.move_in_date,
            "move_out": resident.move_out_date,
        })
    return out


def _to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_fingerprint_sheets(wb, template_id, property_code, template_type, period_start, period_end, roster):
    """Add hidden veryHidden sheets _dt_meta and _dt_roster for template signing (FIX-019 WP-2).

    ``roster`` is a list of dicts: [{resident_id, room_number, bed_code, full_name}, ...]
    These sheets are the ground truth for strict validation + deterministic matching.
    """
    import hashlib, json
    from datetime import datetime as dt

    # _dt_meta: key/value pairs
    ws_meta = wb.create_sheet("_dt_meta")
    ws_meta.sheet_state = "veryHidden"
    roster_json = json.dumps(roster, sort_keys=True, default=str)
    roster_hash = hashlib.sha256(roster_json.encode()).hexdigest()
    meta_rows = [
        ("template_id", str(template_id)),
        ("template_type", template_type),
        ("property_code", property_code),
        ("period_start", str(period_start) if period_start else ""),
        ("period_end", str(period_end) if period_end else ""),
        ("generated_at", dt.utcnow().isoformat()),
        ("generator_version", "FIX-019-v1"),
        ("roster_count", str(len(roster))),
        ("roster_hash", roster_hash),
    ]
    for i, (k, v) in enumerate(meta_rows, start=1):
        ws_meta.cell(row=i, column=1, value=k)
        ws_meta.cell(row=i, column=2, value=v)

    # _dt_roster: one row per resident
    ws_roster = wb.create_sheet("_dt_roster")
    ws_roster.sheet_state = "veryHidden"
    ws_roster.cell(row=1, column=1, value="resident_id")
    ws_roster.cell(row=1, column=2, value="room_number")
    ws_roster.cell(row=1, column=3, value="bed_code")
    ws_roster.cell(row=1, column=4, value="full_name")
    for i, r in enumerate(roster, start=2):
        ws_roster.cell(row=i, column=1, value=r.get("resident_id", ""))
        ws_roster.cell(row=i, column=2, value=r.get("room_number", ""))
        ws_roster.cell(row=i, column=3, value=r.get("bed_code", ""))
        ws_roster.cell(row=i, column=4, value=r.get("full_name", ""))

    return roster_hash


def build_row_template(resident_rows, property_code, prefill_date=None, kind="adhoc", template_id=None, roster=None):
    """Row-per-reading workbook (Ad-hoc or Daily).

    ``prefill_date`` (a ``date``) pre-fills the Reading Date column (Daily). When
    ``None`` (Ad-hoc) the Reading Date cell is left blank + highlighted for input.
    ``template_id`` and ``roster`` (FIX-019 WP-2) embed hidden fingerprint sheets.
    """
    rows = _rows(resident_rows)
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    ws = wb.active
    ws.title = "Meter Readings"

    for i, h in enumerate(STD_HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[1].height = 30

    r = 2
    for row in rows:
        ws.cell(row=r, column=1, value=property_code).alignment = CENTER   # Branch Code
        ws.cell(row=r, column=2, value=property_code).alignment = CENTER   # Building
        ws.cell(row=r, column=3, value=row["room"]).alignment = CENTER     # Room Number
        ws.cell(row=r, column=4, value=row["bed"]).alignment = CENTER      # Bed
        ws.cell(row=r, column=5, value=row["name"]).alignment = LEFT       # Resident Name
        dcell = ws.cell(row=r, column=6, value=prefill_date)               # Reading Date
        dcell.number_format = "yyyy-mm-dd"
        dcell.alignment = CENTER
        if prefill_date is None:
            dcell.fill = FILLIN_FILL
        for col in (7, 8):                                                 # Electric / Water
            cc = ws.cell(row=r, column=col)
            cc.fill = FILLIN_FILL
            cc.number_format = "0.##"
            cc.alignment = CENTER
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    for col, w in zip("ABCDEFGH", [12, 12, 13, 6, 32, 24, 20, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "F2"

    if kind == "daily":
        ws["G1"].comment = Comment(
            "Type the ELECTRIC meter value for each resident (yellow cells). "
            "Add a water reading too if you take one. Rows left blank are skipped on upload.",
            "DormTel")
    else:
        ws["F1"].comment = Comment(
            "Fill in the Reading Date (YYYY-MM-DD) and at least one reading value for each "
            "row you want to upload. Copy a resident's row to record more than one date. "
            "Blank rows are skipped on upload.",
            "DormTel")

    # FIX-019 WP-2: embed hidden fingerprint sheets if template_id provided
    if template_id and roster is not None:
        period_start = prefill_date if kind == "daily" else None
        period_end = prefill_date if kind == "daily" else None
        _add_fingerprint_sheets(wb, template_id, property_code, kind, period_start, period_end, roster)

    return _to_bytes(wb)


def build_monthly_template(resident_rows, property_code, year, month, template_id=None, roster=None):
    """Grid workbook (one column per day of the month) for the Monthly Grid upload.
    ``template_id`` and ``roster`` (FIX-019 WP-2) embed hidden fingerprint sheets.
    """
    rows = _rows(resident_rows)
    ndays = calendar.monthrange(year, month)[1]
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    ws = wb.active
    ws.title = "METER READINGS"

    text_headers = ["ROOM", "BED", "NAME", "RATE", "MOVE IN", "MOVE OUT", "DAYS"]
    for i, h in enumerate(text_headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    first_date_col = len(text_headers) + 1
    for d in range(1, ndays + 1):
        c = ws.cell(row=1, column=first_date_col + (d - 1), value=datetime(year, month, d))
        c.number_format = "dd-mmm"
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    ws["A1"].comment = Comment(
        f"MONTHLY METER GRID — {calendar.month_name[month]} {year} ({property_code})\n"
        "1. Columns A-G are pre-filled from the system.\n"
        "2. Type the DAILY ELECTRIC meter value in the yellow date columns (one per day).\n"
        "3. Leave a date blank if no reading was taken that day.\n"
        "4. DAYS = auto-calculated days the resident was present in the month.\n"
        "5. Upload via Billing > 'Monthly Grid Upload'. Do not rename the date columns.",
        "DormTel")
    ws["A1"].comment.width = 420
    ws["A1"].comment.height = 200

    r = 2
    for row in rows:
        ws.cell(row=r, column=1, value=row["room"]).alignment = CENTER
        ws.cell(row=r, column=2, value=row["bed"]).alignment = CENTER
        ws.cell(row=r, column=3, value=row["name"]).alignment = LEFT
        rc = ws.cell(row=r, column=4, value=row["rate"])
        rc.number_format = "#,##0.00"
        rc.alignment = CENTER
        mi = ws.cell(row=r, column=5, value=row["move_in"])
        mi.number_format = "yyyy-mm-dd"
        mi.alignment = CENTER
        mo = ws.cell(row=r, column=6, value=row["move_out"])
        mo.number_format = "yyyy-mm-dd"
        mo.alignment = CENTER
        days_formula = (
            f'=IF($E{r}="","",'
            f'MAX(0,MIN(IF($F{r}="",DATE({year},{month},{ndays}),$F{r}),DATE({year},{month},{ndays}))'
            f'-MAX($E{r},DATE({year},{month},1))+1))'
        )
        dc = ws.cell(row=r, column=7, value=days_formula)
        dc.number_format = "0"
        dc.alignment = CENTER
        dc.font = BOLD_FONT
        for d in range(ndays):
            fc = ws.cell(row=r, column=first_date_col + d)
            fc.fill = FILLIN_FILL
            fc.alignment = CENTER
            fc.number_format = "0.##"
        for col in range(1, first_date_col + ndays):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 6
    for d in range(ndays):
        ws.column_dimensions[get_column_letter(first_date_col + d)].width = 7
    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 30

    # FIX-019 WP-2: embed hidden fingerprint sheets if template_id provided
    if template_id and roster is not None:
        from datetime import date as _d
        ps = _d(year, month, 1)
        pe = _d(year, month, ndays)
        _add_fingerprint_sheets(wb, template_id, property_code, "monthly", ps, pe, roster)

    return _to_bytes(wb)


def build_period_template(resident_rows, property_code, start_date, end_date, template_id=None, roster=None):
    """Grid workbook (one column per day in the custom period) for Billing Period uploads.

    Similar to ``build_monthly_template`` but spans an arbitrary date range
    (e.g. May 31 – Jun 30). The DAYS formula is clamped to the period range.
    Safety cap: max 62 days.
    ``template_id`` and ``roster`` (FIX-019 WP-2) embed hidden fingerprint sheets.
    """
    from datetime import date, timedelta

    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

    if start_date > end_date:
        raise ValueError("start_date must be before or equal to end_date")

    total_days = (end_date - start_date).days + 1
    if total_days > 62:
        raise ValueError(f"Period too long ({total_days} days). Maximum is 62 days.")

    rows = _rows(resident_rows)
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    ws = wb.active
    ws.title = "METER READINGS"

    text_headers = ["ROOM", "BED", "NAME", "RATE", "MOVE IN", "MOVE OUT", "DAYS"]
    for i, h in enumerate(text_headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    first_date_col = len(text_headers) + 1

    # One date column per day in the period
    for d in range(total_days):
        current_date = start_date + timedelta(days=d)
        c = ws.cell(row=1, column=first_date_col + d, value=datetime(current_date.year, current_date.month, current_date.day))
        c.number_format = "dd-mmm"
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    period_label = f"{start_date.strftime('%b %d')} – {end_date.strftime('%b %d, %Y')}"
    ws["A1"].comment = Comment(
        f"BILLING PERIOD METER GRID — {period_label} ({property_code})\n"
        "1. Columns A-G are pre-filled from the system.\n"
        "2. Type the DAILY ELECTRIC meter value in the yellow date columns.\n"
        "3. Leave a date blank if no reading was taken that day.\n"
        "4. DAYS = auto-calculated days the resident was present in the period.\n"
        "5. Upload via Billing > 'Daily Sheet Upload'. Do not rename the date columns.",
        "DormTel")
    ws["A1"].comment.width = 420
    ws["A1"].comment.height = 200

    r = 2
    for row in rows:
        ws.cell(row=r, column=1, value=row["room"]).alignment = CENTER
        ws.cell(row=r, column=2, value=row["bed"]).alignment = CENTER
        ws.cell(row=r, column=3, value=row["name"]).alignment = LEFT
        rc = ws.cell(row=r, column=4, value=row["rate"])
        rc.number_format = "#,##0.00"
        rc.alignment = CENTER
        mi = ws.cell(row=r, column=5, value=row["move_in"])
        mi.number_format = "yyyy-mm-dd"
        mi.alignment = CENTER
        mo = ws.cell(row=r, column=6, value=row["move_out"])
        mo.number_format = "yyyy-mm-dd"
        mo.alignment = CENTER
        # DAYS formula clamped to the period range
        days_formula = (
            f'=IF($E{r}="","",'
            f'MAX(0,MIN(IF($F{r}="",DATE({end_date.year},{end_date.month},{end_date.day}),$F{r}),DATE({end_date.year},{end_date.month},{end_date.day}))'
            f'-MAX($E{r},DATE({start_date.year},{start_date.month},{start_date.day}))+1))'
        )
        dc = ws.cell(row=r, column=7, value=days_formula)
        dc.number_format = "0"
        dc.alignment = CENTER
        dc.font = BOLD_FONT
        for d in range(total_days):
            fc = ws.cell(row=r, column=first_date_col + d)
            fc.fill = FILLIN_FILL
            fc.alignment = CENTER
            fc.number_format = "0.##"
        for col in range(1, first_date_col + total_days):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 6
    for d in range(total_days):
        ws.column_dimensions[get_column_letter(first_date_col + d)].width = 7
    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 30

    # FIX-019 WP-2: embed hidden fingerprint sheets if template_id provided
    if template_id and roster is not None:
        _add_fingerprint_sheets(wb, template_id, property_code, "period", start_date, end_date, roster)

    return _to_bytes(wb)

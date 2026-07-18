import io
import os
import re
import calendar

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import Optional, List
from decimal import Decimal
from uuid import UUID
from datetime import datetime, date
from pydantic import BaseModel, Field

from app.database import get_db
from app import models, auth
from app.models import MeterReading, MeterReadingImport, Billing, Resident, Room, Bed, Payment, LedgerEntry
from app.schemas import (
    MeterReadingCreate, MeterReadingOut, MeterReadingDailyGridOut,
    MeterReadingDailyRow, MeterReadingDailyCell,
    VacantBedRow, ImportInfo,
    BillingOut, BillingWithResidentOut,
    MeterReadingUploadResult, MeterReadingDailySheetResult,
    BillingImportStatusOut,
)
from app.utils.property_filter import get_property_buildings


class BillingPreviewRow(BaseModel):
    resident_id: UUID
    resident_name: str
    room_number: Optional[str] = None
    bed_number: Optional[int] = None
    bed_code: Optional[str] = None
    monthly_rate: Decimal
    rent_amount: Decimal
    electric_charge: Decimal
    water_charge: Decimal
    other_charges: Decimal
    previous_balance: Optional[Decimal] = Decimal("0")
    total_amount: Decimal


class BillingPreviewResponse(BaseModel):
    billing_period: str
    building: Optional[str] = None
    total_residents: int
    rows: List[BillingPreviewRow]
    summary: dict


def _parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%d/%m/%Y",
                     "%b %d, %Y", "%B %d, %Y", "%d-%b-%Y", "%m-%d-%Y",
                     "%Y/%m/%d", "%d/%m/%y", "%m/%d/%y"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _parse_decimal(val):
    if val is None or val == "":
        return None
    try:
        s = str(val).strip()
        # Remove peso sign, spaces, commas
        s = s.replace("₱", "").replace("PHP", "").replace(" ", "").replace(",", "")
        # Handle parentheses as negative: (1,234) -> -1234
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        if not s or s == "-":
            return None
        return Decimal(s)
    except Exception:
        return None


def _is_floor_marker(val):
    """Detect floor marker rows like '2nd Floor', '3rd Floor', '4F', etc."""
    if val is None:
        return False
    s = str(val).strip().lower()
    if not s:
        return False
    # Match patterns: "2nd floor", "3rd floor", "4f", "ground floor", "basement", etc.
    return bool(re.match(r'^(\d+(st|nd|rd|th)\s*floor?|\d+f|ground\s*floor|basement|mezzanine|penthouse)$', s))


def _normalize_name(name):
    if not name:
        return ""
    # Strip, uppercase, normalize Ñ, collapse all whitespace (spaces, tabs, NBSP)
    normalized = name.strip().upper().replace("\u00d1", "N").replace("\u00f1", "n")
    normalized = re.sub(r'\s+', ' ', normalized)
    return normalized


router = APIRouter()


class BillingGenerateRequest(BaseModel):
    billing_period: str
    building: Optional[str] = None
    other_charges: Optional[Decimal] = Field(default=None, ge=0)
    total_water_bill: Optional[Decimal] = Field(default=None, ge=0)


async def _get_active_residents_with_rooms(
    db: AsyncSession,
    building: Optional[str] = None,
    property_code: Optional[str] = None,
):
    """Get active residents joined with their bed and room."""
    query = (
        select(Resident, Bed, Room)
        .join(Bed, Resident.bed_id == Bed.id, isouter=True)
        .join(Room, Bed.room_id == Room.id, isouter=True)
        .where(Resident.status == "active")
    )
    if property_code:
        query = query.where(Room.property_code == property_code)
    if building:
        query = query.where(Room.building == building)

    result = await db.execute(query)
    rows = result.all()

    # Group by room
    residents_by_room = {}
    no_room = []
    for resident, bed, room in rows:
        if room:
            rid = str(room.id)
            if rid not in residents_by_room:
                residents_by_room[rid] = {"room": room, "residents": []}
            residents_by_room[rid]["residents"].append((resident, bed))
        else:
            no_room.append((resident, bed))

    return residents_by_room, no_room, rows


async def _compute_resident_electric(
    db: AsyncSession,
    billing_period: str,
    building: Optional[str] = None,
    property_code: Optional[str] = None,
):
    """Sum per-resident electric meter readings for the billing period.
    If a daily-sheet import exists for the resident, use the pre-computed total."""
    year, month = map(int, billing_period.split("-"))
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    # Resolve building filter from property_code if provided
    property_buildings = None
    if property_code:
        property_buildings = await get_property_buildings(db, property_code)

    # Check for imported daily-sheet totals first
    import_query = select(MeterReadingImport)
    if property_buildings:
        import_query = import_query.where(MeterReadingImport.building.in_(property_buildings))
    elif building:
        import_query = import_query.where(MeterReadingImport.building == building)
    import_query = import_query.where(
        MeterReadingImport.year == year,
        MeterReadingImport.month == month,
    )
    import_result = await db.execute(import_query)
    imports = import_result.scalars().all()

    resident_electric = {}
    imported_residents = set()
    for imp in imports:
        rid = str(imp.resident_id)
        if imp.total_electric_usage is not None:
            resident_electric[rid] = imp.total_electric_usage
            imported_residents.add(rid)

    # For residents without an import, fall back to computing usage from cumulative readings
    query = select(MeterReading).where(
        MeterReading.reading_date >= start_date,
        MeterReading.reading_date < end_date,
    )
    if property_buildings:
        query = query.where(MeterReading.building.in_(property_buildings))
    elif building:
        query = query.where(MeterReading.building == building)

    result = await db.execute(query)
    readings = result.scalars().all()

    resident_readings = {}
    for reading in readings:
        rid = str(reading.resident_id)
        if rid in imported_residents:
            continue
        if rid not in resident_readings:
            resident_readings[rid] = []
        resident_readings[rid].append(reading.electric_reading or Decimal("0"))

    for rid, values in resident_readings.items():
        if len(values) >= 2:
            resident_electric[rid] = max(values) - min(values)
        elif len(values) == 1:
            resident_electric[rid] = values[0]

    return resident_electric


def _count_days_stayed(move_in_date, move_out_date, period_start, period_end):
    """Count overlap days between resident stay and billing period."""
    if not move_in_date:
        # No move-in date: assume full period
        effective_start = period_start
    else:
        effective_start = max(move_in_date, period_start)

    if move_out_date and move_out_date < period_start:
        return 0

    if move_out_date:
        effective_end = min(move_out_date, period_end)
    else:
        effective_end = period_end

    if effective_start > effective_end:
        return 0

    return (effective_end - effective_start).days


async def _compute_water_by_days(
    db: AsyncSession,
    billing_period: str,
    total_water_bill: Optional[Decimal] = None,
    building: Optional[str] = None,
    property_code: Optional[str] = None,
):
    """Compute water charge per resident based on days stayed.
    If a daily-sheet import exists with a water_bill, use it directly."""
    year, month = map(int, billing_period.split("-"))
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    query = (
        select(Resident, Bed, Room)
        .join(Bed, Resident.bed_id == Bed.id, isouter=True)
        .join(Room, Bed.room_id == Room.id, isouter=True)
        .where(Resident.status == "active")
    )
    if property_code:
        query = query.where(Room.property_code == property_code)
    if building:
        query = query.where(Room.building == building)

    result = await db.execute(query)
    rows = result.all()

    # Resolve building filter from property_code if provided
    property_buildings = None
    if property_code:
        property_buildings = await get_property_buildings(db, property_code)

    # Check for imported water bills
    import_query = select(MeterReadingImport)
    if property_buildings:
        import_query = import_query.where(MeterReadingImport.building.in_(property_buildings))
    elif building:
        import_query = import_query.where(MeterReadingImport.building == building)
    import_query = import_query.where(
        MeterReadingImport.year == year,
        MeterReadingImport.month == month,
    )
    import_result = await db.execute(import_query)
    imports = import_result.scalars().all()

    imported_water = {str(imp.resident_id): imp.water_bill for imp in imports if imp.water_bill is not None}

    if imported_water:
        # Use imported water bills directly; return dummy total_days/rate
        resident_water = {}
        for resident, bed, room in rows:
            rid = str(resident.id)
            resident_water[rid] = imported_water.get(rid, Decimal("0"))
        return resident_water, 0, Decimal("0")

    total_days = 0
    resident_days = {}
    for resident, bed, room in rows:
        days = _count_days_stayed(resident.move_in_date, resident.move_out_date, start_date, end_date)
        total_days += days
        resident_days[str(resident.id)] = days

    water_input = total_water_bill or Decimal("0")
    rate_per_day = water_input / total_days if total_days > 0 else Decimal("0")

    resident_water = {}
    for rid, days in resident_days.items():
        resident_water[rid] = Decimal(str(days)) * rate_per_day

    return resident_water, total_days, rate_per_day


async def _compute_other_charges(
    db: AsyncSession,
    billing_period: str,
    other_charges_input: Optional[Decimal] = None,
    building: Optional[str] = None,
    property_code: Optional[str] = None,
):
    """Compute other charges per resident.
    If imports exist with misc charges, use those per resident.
    Otherwise divide total_other_charges equally."""
    year, month = map(int, billing_period.split("-"))

    # Get resident count
    query = (
        select(Resident)
        .join(Bed, Resident.bed_id == Bed.id, isouter=True)
        .join(Room, Bed.room_id == Room.id, isouter=True)
        .where(Resident.status == "active")
    )
    if property_code:
        query = query.where(Room.property_code == property_code)
    if building:
        query = query.where(Room.building == building)
    result = await db.execute(query)
    residents = result.scalars().all()
    total_residents = len(residents)

    # Resolve building filter from property_code if provided
    property_buildings = None
    if property_code:
        property_buildings = await get_property_buildings(db, property_code)

    # Check for imported misc charges
    import_query = select(MeterReadingImport)
    if property_buildings:
        import_query = import_query.where(MeterReadingImport.building.in_(property_buildings))
    elif building:
        import_query = import_query.where(MeterReadingImport.building == building)
    import_query = import_query.where(
        MeterReadingImport.year == year,
        MeterReadingImport.month == month,
    )
    import_result = await db.execute(import_query)
    imports = import_result.scalars().all()

    imported_misc = {}
    if imports:
        for imp in imports:
            if imp.misc_charges:
                total_misc = Decimal("0")
                for key, val in imp.misc_charges.items():
                    try:
                        total_misc += Decimal(str(val))
                    except Exception:
                        pass
                if total_misc > 0:
                    imported_misc[str(imp.resident_id)] = total_misc

    if imported_misc:
        result_map = {}
        for resident in residents:
            result_map[str(resident.id)] = imported_misc.get(str(resident.id), Decimal("0"))
        return result_map, Decimal("0")

    # Fallback: divide equally
    total_other = other_charges_input or Decimal("0")
    per_head = total_other / total_residents if total_residents > 0 else Decimal("0")
    result_map = {}
    for resident in residents:
        result_map[str(resident.id)] = per_head
    return result_map, total_other


async def _create_debit_ledger_entry(
    db: AsyncSession,
    resident_id,
    amount: Decimal,
    description: str,
    reference_id,
):
    """Create a debit ledger entry for a billing record."""
    # Get last running balance
    result = await db.execute(
        select(LedgerEntry)
        .where(LedgerEntry.resident_id == resident_id)
        .order_by(LedgerEntry.created_at.desc())
        .limit(1)
    )
    last_entry = result.scalar_one_or_none()
    running_balance = last_entry.running_balance if last_entry else Decimal("0.00")
    new_balance = running_balance + amount

    ledger = LedgerEntry(
        resident_id=resident_id,
        entry_type="debit",
        amount=amount,
        description=description,
        reference_id=reference_id,
        running_balance=new_balance,
    )
    db.add(ledger)


async def _compute_previous_balances(
    db: AsyncSession,
    billing_period: str,
    building: Optional[str] = None,
    property_code: Optional[str] = None,
) -> dict:
    """Compute unpaid balances from prior billing periods for each resident."""
    year, month = map(int, billing_period.split("-"))
    # Previous period
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1
    prev_period = f"{prev_year}-{prev_month:02d}"

    # Find billings from previous period
    query = (
        select(Billing, Resident)
        .join(Resident, Billing.resident_id == Resident.id)
        .where(Billing.billing_period == prev_period)
    )
    if building or property_code:
        query = query.join(Bed, Resident.bed_id == Bed.id, isouter=True)
        query = query.join(Room, Bed.room_id == Room.id, isouter=True)
        if property_code:
            query = query.where(Room.property_code == property_code)
        if building:
            query = query.where(Room.building == building)

    result = await db.execute(query)
    rows = result.all()

    balances = {}
    for billing, resident in rows:
        rid = str(resident.id)
        # Check payments against this billing
        payment_query = select(Payment).where(
            Payment.billing_id == billing.id,
            Payment.status.in_(["matched", "verified"]),
        )
        payment_result = await db.execute(payment_query)
        payments = payment_result.scalars().all()
        total_paid = sum((p.amount or Decimal("0")) for p in payments)
        unpaid = billing.total_amount - total_paid
        if unpaid > 0:
            balances[rid] = unpaid

    return balances


@router.post("/meter-readings", response_model=MeterReadingOut)
async def submit_meter_reading(
    data: MeterReadingCreate,
    current_staff: models.Staff = Depends(auth.require_staff),
    db: AsyncSession = Depends(get_db),
):
    # Validate room if provided
    room = None
    if data.room_id:
        room_result = await db.execute(select(Room).where(Room.id == data.room_id))
        room = room_result.scalar_one_or_none()
        if not room:
            raise HTTPException(status_code=400, detail="Room not found")

    # Validate resident if provided
    resident = None
    if data.resident_id:
        resident_result = await db.execute(select(Resident).where(Resident.id == data.resident_id))
        resident = resident_result.scalar_one_or_none()
        if not resident:
            raise HTTPException(status_code=400, detail="Resident not found")

    # Previous reading for variance (same resident if resident_id provided, else same room/building)
    prev_query = select(MeterReading).where(
        MeterReading.building == data.building,
        MeterReading.reading_date < data.reading_date
    )
    if data.resident_id:
        prev_query = prev_query.where(MeterReading.resident_id == data.resident_id)
    elif data.room_id:
        prev_query = prev_query.where(MeterReading.room_id == data.room_id)
    prev_query = prev_query.order_by(MeterReading.reading_date.desc()).limit(1)

    prev_result = await db.execute(prev_query)
    prev = prev_result.scalar_one_or_none()

    variance_pct = Decimal("0")
    if prev:
        prev_electric = prev.electric_reading or Decimal("0")
        prev_water = prev.water_reading or Decimal("0")
        curr_electric = data.electric_reading or Decimal("0")
        curr_water = data.water_reading or Decimal("0")
        prev_total = prev_electric + prev_water
        curr_total = curr_electric + curr_water
        if prev_total > 0:
            variance_pct = ((curr_total - prev_total) / prev_total) * 100

    status = "pending" if variance_pct > 20 else "approved"

    reading = MeterReading(
        building=data.building,
        room_id=data.room_id,
        resident_id=data.resident_id,
        reading_date=data.reading_date,
        electric_reading=data.electric_reading,
        water_reading=data.water_reading,
        variance_pct=variance_pct,
        status=status,
    )
    db.add(reading)
    await db.commit()
    await db.refresh(reading)

    # Attach related data for response
    bed = None
    if resident and resident.bed_id:
        bed_result = await db.execute(select(Bed).where(Bed.id == resident.bed_id))
        bed = bed_result.scalar_one_or_none()

    out = MeterReadingOut(
        id=reading.id,
        building=reading.building,
        room_id=reading.room_id,
        resident_id=reading.resident_id,
        reading_date=reading.reading_date,
        electric_reading=reading.electric_reading,
        water_reading=reading.water_reading,
        status=reading.status,
        variance_pct=reading.variance_pct,
        room_number=room.room_number if room else None,
        resident_name=resident.full_name if resident else None,
        bed_code=bed.bed_code if bed else None,
    )
    return out


@router.post("/generate", response_model=List[BillingOut])
async def generate_billings(
    data: BillingGenerateRequest,
    current_staff: models.Staff = Depends(auth.require_staff),
    property_code: Optional[str] = Depends(auth.get_current_property),
    db: AsyncSession = Depends(get_db),
):
    residents_by_room, no_room, all_rows = await _get_active_residents_with_rooms(db, data.building, property_code)
    if not all_rows:
        detail = f"No active residents found" + (f" for building '{data.building}'" if data.building else "")
        raise HTTPException(status_code=400, detail=detail)

    # Idempotency: do not duplicate billings for the same period
    resident_ids = [str(resident.id) for resident, bed, room in all_rows]
    existing_result = await db.execute(
        select(Billing).where(
            Billing.billing_period == data.billing_period,
            Billing.resident_id.in_(resident_ids),
        )
    )
    existing = existing_result.scalars().all()
    if existing:
        return existing

    resident_electric = await _compute_resident_electric(db, data.billing_period, data.building, property_code)
    resident_water, total_days, rate_per_day = await _compute_water_by_days(
        db, data.billing_period, data.total_water_bill, data.building, property_code
    )
    resident_other, total_other = await _compute_other_charges(
        db, data.billing_period, data.other_charges, data.building, property_code
    )
    previous_balances = await _compute_previous_balances(db, data.billing_period, data.building, property_code)

    billings = []

    # Residents with rooms
    for room_id, room_data in residents_by_room.items():
        room_residents = room_data["residents"]

        for resident, bed in room_residents:
            rid = str(resident.id)
            elec = resident_electric.get(rid, Decimal("0"))
            wat = resident_water.get(rid, Decimal("0"))
            other = resident_other.get(rid, Decimal("0"))
            prev_bal = previous_balances.get(rid, Decimal("0"))
            total = resident.monthly_rate + elec + wat + other + prev_bal
            billing = Billing(
                resident_id=resident.id,
                billing_period=data.billing_period,
                rent_amount=resident.monthly_rate,
                electric_charge=elec,
                water_charge=wat,
                other_charges=other,
                previous_balance=prev_bal,
                total_amount=total,
                variance_pct=Decimal("0"),
                status="draft",
            )
            db.add(billing)
            billings.append(billing)

    # Residents without a room (fallback: rent + other only, no utilities)
    for resident, bed in no_room:
        rid = str(resident.id)
        elec = resident_electric.get(rid, Decimal("0"))
        wat = resident_water.get(rid, Decimal("0"))
        other = resident_other.get(rid, Decimal("0"))
        prev_bal = previous_balances.get(rid, Decimal("0"))
        total = resident.monthly_rate + elec + wat + other + prev_bal
        billing = Billing(
            resident_id=resident.id,
            billing_period=data.billing_period,
            rent_amount=resident.monthly_rate,
            electric_charge=elec,
            water_charge=wat,
            other_charges=other,
            previous_balance=prev_bal,
            total_amount=total,
            variance_pct=Decimal("0"),
            status="draft",
        )
        db.add(billing)
        billings.append(billing)

    await db.flush()

    # Create debit ledger entries for each billing
    for billing in billings:
        await _create_debit_ledger_entry(
            db,
            resident_id=billing.resident_id,
            amount=billing.total_amount,
            description=f"Billing for {billing.billing_period}",
            reference_id=billing.id,
        )

    await db.commit()
    for b in billings:
        await db.refresh(b)
    return billings


@router.post("/{billing_id}/approve", response_model=BillingOut)
async def approve_billing(
    billing_id: UUID,
    current_staff: models.Staff = Depends(auth.require_staff),
    db: AsyncSession = Depends(get_db),
):
    billing = await db.get(Billing, billing_id)
    if not billing:
        raise HTTPException(status_code=404, detail="Billing not found")
    billing.status = "approved"
    await db.commit()
    await db.refresh(billing)
    return billing


@router.post("/{billing_id}/distribute", response_model=BillingOut)
async def distribute_billing(
    billing_id: UUID,
    current_staff: models.Staff = Depends(auth.require_staff),
    db: AsyncSession = Depends(get_db),
):
    billing = await db.get(Billing, billing_id)
    if not billing:
        raise HTTPException(status_code=404, detail="Billing not found")
    billing.status = "distributed"
    billing.distributed_at = datetime.utcnow()
    billing.payment_link = f"https://pay.dormtel.app/b/{billing_id}"
    await db.commit()
    await db.refresh(billing)
    return billing


# ─── Template Validation (pre-upload structural check) ───

def _validate_standard_template(contents: bytes, filename: str):
    """Validate standard meter reading template structure."""
    from openpyxl import load_workbook
    issues = []
    sheets_preview = []

    if not filename.lower().endswith((".xlsx", ".xls")):
        issues.append(schemas.TemplateValidationIssue(
            severity="error", code="INVALID_FILE_TYPE",
            message=f"File '{filename}' is not an Excel file. Please save as .xlsx format."))
        return schemas.TemplateValidationResponse(
            upload_type="standard", file_name=filename, file_size_bytes=len(contents),
            overall_status="invalid", issues=issues, sheets=[], summary={})

    if len(contents) > 10 * 1024 * 1024:
        issues.append(schemas.TemplateValidationIssue(
            severity="error", code="FILE_TOO_LARGE",
            message=f"File is {len(contents)/1024/1024:.1f} MB. Maximum allowed is 10 MB."))
        return schemas.TemplateValidationResponse(
            upload_type="standard", file_name=filename, file_size_bytes=len(contents),
            overall_status="invalid", issues=issues, sheets=[], summary={})
    if len(contents) < 100:
        issues.append(schemas.TemplateValidationIssue(
            severity="error", code="FILE_EMPTY",
            message="File appears to be empty or corrupted."))
        return schemas.TemplateValidationResponse(
            upload_type="standard", file_name=filename, file_size_bytes=len(contents),
            overall_status="invalid", issues=issues, sheets=[], summary={})

    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True, keep_links=False, read_only=True)
    except Exception as e:
        error_msg = str(e)
        if "not a zip file" in error_msg.lower():
            issues.append(schemas.TemplateValidationIssue(
                severity="error", code="FILE_NOT_PARSEABLE",
                message="File format not recognized. If using Google Sheets, use File > Download > Microsoft Excel (.xlsx)."))
        elif "password" in error_msg.lower():
            issues.append(schemas.TemplateValidationIssue(
                severity="error", code="PASSWORD_PROTECTED",
                message="File is password-protected. Please remove the password before uploading."))
        else:
            issues.append(schemas.TemplateValidationIssue(
                severity="error", code="FILE_NOT_PARSEABLE",
                message=f"Cannot read this file: {error_msg[:200]}"))
        return schemas.TemplateValidationResponse(
            upload_type="standard", file_name=filename, file_size_bytes=len(contents),
            overall_status="invalid", issues=issues, sheets=[], summary={})

    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == "meter readings":
            ws = wb[name]
            break
    if ws is None:
        available = ", ".join(wb.sheetnames[:5])
        issues.append(schemas.TemplateValidationIssue(
            severity="error", code="MISSING_SHEET",
            message=f"Sheet 'Meter Readings' not found. Found: [{available}]. Please rename your data sheet or download the template."))
        wb.close()
        return schemas.TemplateValidationResponse(
            upload_type="standard", file_name=filename, file_size_bytes=len(contents),
            overall_status="invalid", issues=issues, sheets=[], summary={})

    REQUIRED_COLS = {"Branch Code", "Building", "Room Number", "Bed", "Resident Name",
                     "Reading Date (YYYY-MM-DD)", "Electric Reading (kWh)", "Water Reading (m³)"}
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")
    header_set = {h for h in headers if h}
    detected = [h for h in headers if h]
    missing = sorted(REQUIRED_COLS - header_set)
    extra = sorted(header_set - REQUIRED_COLS)

    for col_name in missing:
        issues.append(schemas.TemplateValidationIssue(
            severity="error", code="MISSING_REQUIRED_COLUMN",
            message=f"Column '{col_name}' is missing. This column is required.",
            sheet="Meter Readings", column=col_name))
    if extra:
        issues.append(schemas.TemplateValidationIssue(
            severity="info", code="EXTRA_COLUMNS",
            message=f"{len(extra)} extra column(s) will be ignored: [{', '.join(extra[:5])}]",
            sheet="Meter Readings"))

    data_row_count = 0
    sample_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            data_row_count += 1
            if len(sample_rows) < 3:
                sample_rows.append([str(v) if v is not None else "" for v in row[:10]])

    if data_row_count == 0:
        issues.append(schemas.TemplateValidationIssue(
            severity="warning", code="EMPTY_DATA_ROWS",
            message="Sheet has headers but no data rows.", sheet="Meter Readings"))
    elif data_row_count < 3:
        issues.append(schemas.TemplateValidationIssue(
            severity="info", code="FEW_DATA_ROWS",
            message=f"Only {data_row_count} data row(s) found.", sheet="Meter Readings"))

    sheets_preview.append(schemas.SheetPreview(
        name="Meter Readings", header_row_index=1,
        detected_headers=detected, missing_headers=missing, extra_headers=extra,
        data_row_count=data_row_count, sample_rows=sample_rows))

    wb.close()
    has_errors = any(i.severity == "error" for i in issues)
    has_warnings = any(i.severity == "warning" for i in issues)
    status = "invalid" if has_errors else ("warnings" if has_warnings else "valid")
    return schemas.TemplateValidationResponse(
        upload_type="standard", file_name=filename, file_size_bytes=len(contents),
        overall_status=status, issues=issues, sheets=sheets_preview,
        summary={"data_rows": data_row_count, "sheets": 1})


def _validate_daily_sheet_template(contents: bytes, filename: str):
    """Validate daily sheet template structure."""
    from openpyxl import load_workbook
    from collections import Counter
    issues = []
    sheets_preview = []

    if not filename.lower().endswith((".xlsx", ".xls")):
        issues.append(schemas.TemplateValidationIssue(
            severity="error", code="INVALID_FILE_TYPE",
            message=f"File '{filename}' is not an Excel file. Please save as .xlsx format."))
        return schemas.TemplateValidationResponse(
            upload_type="daily_sheet", file_name=filename, file_size_bytes=len(contents),
            overall_status="invalid", issues=issues, sheets=[], summary={})
    if len(contents) > 10 * 1024 * 1024:
        issues.append(schemas.TemplateValidationIssue(
            severity="error", code="FILE_TOO_LARGE",
            message=f"File is {len(contents)/1024/1024:.1f} MB. Maximum allowed is 10 MB."))
        return schemas.TemplateValidationResponse(
            upload_type="daily_sheet", file_name=filename, file_size_bytes=len(contents),
            overall_status="invalid", issues=issues, sheets=[], summary={})
    if len(contents) < 100:
        issues.append(schemas.TemplateValidationIssue(
            severity="error", code="FILE_EMPTY",
            message="File appears to be empty or corrupted."))
        return schemas.TemplateValidationResponse(
            upload_type="daily_sheet", file_name=filename, file_size_bytes=len(contents),
            overall_status="invalid", issues=issues, sheets=[], summary={})

    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True, keep_links=False, read_only=True)
    except Exception as e:
        error_msg = str(e)
        if "not a zip file" in error_msg.lower():
            issues.append(schemas.TemplateValidationIssue(
                severity="error", code="FILE_NOT_PARSEABLE",
                message="File format not recognized. If using Google Sheets, use File > Download > Microsoft Excel (.xlsx)."))
        elif "password" in error_msg.lower():
            issues.append(schemas.TemplateValidationIssue(
                severity="error", code="PASSWORD_PROTECTED",
                message="File is password-protected. Please remove the password before uploading."))
        else:
            issues.append(schemas.TemplateValidationIssue(
                severity="error", code="FILE_NOT_PARSEABLE",
                message=f"Cannot read this file: {error_msg[:200]}"))
        return schemas.TemplateValidationResponse(
            upload_type="daily_sheet", file_name=filename, file_size_bytes=len(contents),
            overall_status="invalid", issues=issues, sheets=[], summary={})

    if not wb.sheetnames:
        issues.append(schemas.TemplateValidationIssue(
            severity="error", code="NO_SHEETS",
            message="Workbook contains no sheets."))
        wb.close()
        return schemas.TemplateValidationResponse(
            upload_type="daily_sheet", file_name=filename, file_size_bytes=len(contents),
            overall_status="invalid", issues=issues, sheets=[], summary={})

    if len(wb.sheetnames) > 1:
        issues.append(schemas.TemplateValidationIssue(
            severity="info", code="MULTIPLE_SHEETS_INFO",
            message=f"{len(wb.sheetnames)} building sheet(s) detected: {', '.join(wb.sheetnames[:10])}"))

    total_residents = 0
    all_sheet_names = list(wb.sheetnames)
    for sheet_name in all_sheet_names[:10]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
        if not rows:
            issues.append(schemas.TemplateValidationIssue(
                severity="error", code="EMPTY_SHEET",
                message=f"Sheet '{sheet_name}' is empty.", sheet=sheet_name))
            continue

        header_row_idx = None
        header_row = None
        max_dates = 0
        for r_idx, row_vals in enumerate(rows):
            date_count = sum(1 for v in row_vals if isinstance(v, datetime))
            if date_count > max_dates:
                max_dates = date_count
                header_row_idx = r_idx + 1
                header_row = row_vals

        if not header_row or max_dates == 0:
            issues.append(schemas.TemplateValidationIssue(
                severity="error", code="NO_DATE_COLUMNS",
                message=f"No date columns found in sheet '{sheet_name}'. Daily sheets need date columns in the header row.",
                sheet=sheet_name))
            continue

        date_cols = []
        header_map = {}
        col_total_usage = False
        col_water_bill = False
        misc_cols = []
        bed_header_idx = None

        for idx, val in enumerate(header_row):
            if val is None:
                continue
            if isinstance(val, datetime):
                date_cols.append(val)
            elif isinstance(val, str):
                parsed = _parse_header_date(val)
                if parsed:
                    date_cols.append(parsed)
                else:
                    vup = val.upper().strip()
                    header_map[vup] = idx
                    if vup == "BED":
                        bed_header_idx = idx
                    elif vup == "TOTAL USAGE":
                        col_total_usage = True
                    elif vup == "WATER BILL":
                        col_water_bill = True
                    elif vup in {"LAUNDRY", "DRINKING WATER", "ICE CREAM", "HONESTY STORE", "COFFEE", "LOST KEYCARD", "REF RENTAL"}:
                        misc_cols.append(vup)

        date_range_start = min(date_cols).strftime("%Y-%m-%d") if date_cols else None
        date_range_end = max(date_cols).strftime("%Y-%m-%d") if date_cols else None
        month_counts = Counter(d.month for d in date_cols)
        year_counts = Counter(d.year for d in date_cols)
        detected_month = calendar.month_name[month_counts.most_common(1)[0][0]] if month_counts else None
        detected_year = year_counts.most_common(1)[0][0] if year_counts else None

        has_bed = bed_header_idx is not None
        format_variant = "new_with_flag" if has_bed and bed_header_idx == 2 else ("standard" if has_bed else "no_bed_header")

        data_row_count = 0
        sample_rows = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or all(v is None for v in row):
                continue
            if _is_floor_marker(row[0]):
                continue
            if row[0] is not None and isinstance(row[0], str) and row[0].strip().lower() in ("total", "totals", "grand total", "summary"):
                continue
            data_row_count += 1
            if len(sample_rows) < 3:
                sample_rows.append([str(v) if v is not None else "" for v in row[:10]])

        total_residents += data_row_count

        if data_row_count == 0:
            issues.append(schemas.TemplateValidationIssue(
                severity="warning", code="EMPTY_DATA_ROWS",
                message=f"Sheet '{sheet_name}' has headers but no data rows.", sheet=sheet_name))

        detected_headers = sorted(header_map.keys())
        sheets_preview.append(schemas.SheetPreview(
            name=sheet_name, header_row_index=header_row_idx,
            detected_headers=detected_headers,
            data_row_count=data_row_count, sample_rows=sample_rows,
            date_column_count=len(date_cols),
            date_range_start=date_range_start, date_range_end=date_range_end,
            detected_month=detected_month, detected_year=detected_year,
            has_bed_column=has_bed, format_variant=format_variant,
            misc_columns=misc_cols,
            has_total_usage=col_total_usage, has_water_bill=col_water_bill))

    wb.close()
    has_errors = any(i.severity == "error" for i in issues)
    has_warnings = any(i.severity == "warning" for i in issues)
    status = "invalid" if has_errors else ("warnings" if has_warnings else "valid")
    return schemas.TemplateValidationResponse(
        upload_type="daily_sheet", file_name=filename, file_size_bytes=len(contents),
        overall_status=status, issues=issues, sheets=sheets_preview,
        summary={"sheets": len(all_sheet_names), "total_residents": total_residents,
                 "previewed_sheets": len(sheets_preview)})


@router.post("/meter-readings/validate-template", response_model=schemas.TemplateValidationResponse)
async def validate_meter_reading_template(
    file: UploadFile = File(...),
    current_staff: models.Staff = Depends(auth.require_staff),
):
    """Lightweight structural check before standard upload. No DB access."""
    contents = await file.read()
    return _validate_standard_template(contents, file.filename)


@router.post("/meter-readings/validate-daily-sheet", response_model=schemas.TemplateValidationResponse)
async def validate_daily_sheet_template(
    file: UploadFile = File(...),
    current_staff: models.Staff = Depends(auth.require_staff),
):
    """Lightweight structural check before daily sheet upload. No DB access."""
    contents = await file.read()
    return _validate_daily_sheet_template(contents, file.filename)


@router.get("/meter-readings/template")
async def download_meter_reading_template(
    current_staff: models.Staff = Depends(auth.require_staff),
):
    template_path = os.path.join(os.path.dirname(__file__), "..", "templates", "meter_reading_template.xlsx")
    if not os.path.exists(template_path):
        raise HTTPException(status_code=404, detail="Template not found")
    with open(template_path, "rb") as f:
        data = f.read()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=meter_reading_template.xlsx"}
    )


@router.post("/meter-readings/upload", response_model=MeterReadingUploadResult)
async def upload_meter_readings(
    file: UploadFile = File(...),
    current_staff: models.Staff = Depends(auth.require_staff),
    property_code: Optional[str] = Depends(auth.get_current_property),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")

    # File size validation (max 10MB)
    MAX_FILE_SIZE = 10 * 1024 * 1024
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"File too large ({len(contents) / 1024 / 1024:.1f}MB). Maximum size is 10MB.")
    if len(contents) < 100:
        raise HTTPException(status_code=400, detail="File appears to be empty or corrupted.")

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(contents), data_only=True, keep_links=False)
        # Look for "Meter Readings" sheet (case-insensitive)
        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() == "meter readings":
                ws = wb[name]
                break
        if ws is None:
            available = ", ".join(wb.sheetnames[:5])
            raise HTTPException(
                status_code=400,
                detail=f"Sheet 'Meter Readings' not found. Available sheets: {available}. "
                       f"Please use the Template button to download the correct format."
            )
    except HTTPException:
        raise
    except Exception as e:
        error_msg = str(e)
        if "not a zip file" in error_msg.lower() or "file format" in error_msg.lower():
            raise HTTPException(
                status_code=400,
                detail="File format not recognized. Please save as .xlsx (Excel Workbook) format. "
                       "If using Google Sheets, use File → Download → Microsoft Excel (.xlsx)."
            )
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {error_msg}")

    # Read headers from first row
    headers = [cell.value for cell in ws[1]]
    header_map = {h: i for i, h in enumerate(headers) if h}

    required_cols = {"Branch Code", "Building", "Room Number", "Bed", "Resident Name",
                     "Reading Date (YYYY-MM-DD)", "Electric Reading (kWh)", "Water Reading (m³)"}
    missing = required_cols - set(header_map.keys())
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

    # Pre-load rooms and residents for lookup (scoped to property)
    room_query = select(Room)
    if property_code:
        room_query = room_query.where(Room.property_code == property_code)
    rooms_result = await db.execute(room_query)
    rooms = rooms_result.scalars().all()
    room_by_number = {r.room_number.strip().upper(): r for r in rooms}

    resident_query = (
        select(Resident, Bed, Room)
        .join(Bed, Resident.bed_id == Bed.id, isouter=True)
        .join(Room, Bed.room_id == Room.id, isouter=True)
        .where(Resident.status == "active")
    )
    if property_code:
        resident_query = resident_query.where(
            (Room.property_code == property_code) | (Room.id.is_(None))
        )
    residents_result = await db.execute(resident_query)
    residents_rows = residents_result.all()
    resident_lookup = {}
    short_bed_lookup = {}
    for resident, bed, room in residents_rows:
        key = _normalize_name(resident.full_name)
        has_bed = bed is not None and bed.bed_code is not None
        if key and key not in resident_lookup:
            resident_lookup[key] = resident
        elif key and has_bed:
            resident_lookup[key] = resident
        if has_bed:
            resident_lookup[bed.bed_code.upper()] = resident
            if room:
                short_code = f"{room.room_number.strip().upper()}{bed.bed_code[-1].upper()}"
                short_bed_lookup[short_code] = resident

    imported = 0
    skipped = 0
    errors = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        building = str(row[header_map.get("Building")]).strip() if header_map.get("Building") is not None and row[header_map.get("Building")] is not None else None
        room_number = str(row[header_map.get("Room Number")]).strip() if header_map.get("Room Number") is not None and row[header_map.get("Room Number")] is not None else None
        resident_name = str(row[header_map.get("Resident Name")]).strip() if header_map.get("Resident Name") is not None and row[header_map.get("Resident Name")] is not None else None
        reading_date = _parse_date(row[header_map.get("Reading Date (YYYY-MM-DD)")])
        electric = _parse_decimal(row[header_map.get("Electric Reading (kWh)")])
        water = _parse_decimal(row[header_map.get("Water Reading (m³)")])

        if not building or not reading_date:
            skipped += 1
            continue
        if electric is None and water is None:
            skipped += 1
            continue

        room_id = None
        if room_number:
            room = room_by_number.get(room_number.upper())
            if room:
                room_id = room.id
            else:
                errors.append(f"Room '{room_number}' not found")

        resident_id = None
        if resident_name:
            norm_name = _normalize_name(resident_name)
            resident = resident_lookup.get(norm_name)
            # Fallback: try short bed code lookup
            if not resident and room_number and header_map.get("Bed") is not None:
                bed_letter = str(row[header_map.get("Bed")]).strip() if row[header_map.get("Bed")] is not None else None
                if bed_letter:
                    short_code = f"{room_number.strip().upper()}{bed_letter.strip().upper()}"
                    resident = short_bed_lookup.get(short_code)
            if resident:
                resident_id = resident.id

        prev_query = select(MeterReading).where(
            MeterReading.building == building,
            MeterReading.reading_date < reading_date
        )
        if resident_id:
            prev_query = prev_query.where(MeterReading.resident_id == resident_id)
        elif room_id:
            prev_query = prev_query.where(MeterReading.room_id == room_id)
        prev_query = prev_query.order_by(MeterReading.reading_date.desc()).limit(1)

        prev_result = await db.execute(prev_query)
        prev = prev_result.scalar_one_or_none()

        variance_pct = Decimal("0")
        if prev:
            prev_electric = prev.electric_reading or Decimal("0")
            prev_water = prev.water_reading or Decimal("0")
            curr_electric = electric or Decimal("0")
            curr_water = water or Decimal("0")
            prev_total = prev_electric + prev_water
            curr_total = curr_electric + curr_water
            if prev_total > 0:
                variance_pct = ((curr_total - prev_total) / prev_total) * 100

        status = "pending" if variance_pct > 20 else "approved"

        reading = MeterReading(
            building=building,
            room_id=room_id,
            resident_id=resident_id,
            reading_date=reading_date,
            electric_reading=electric,
            water_reading=water,
            variance_pct=variance_pct,
            status=status,
        )
        db.add(reading)
        imported += 1

    await db.commit()
    return MeterReadingUploadResult(
        imported=imported,
        skipped=skipped,
        errors=errors,
        message=f"Imported {imported} meter readings. Skipped {skipped} rows."
    )


def _parse_header_date(val: str):
    """Try to parse a string as a date using common formats."""
    from datetime import datetime
    formats = ["%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]
    for fmt in formats:
        try:
            return datetime.strptime(val.strip(), fmt)
        except ValueError:
            continue
    return None


@router.post("/meter-readings/upload-daily-sheet", response_model=MeterReadingDailySheetResult)
async def upload_daily_meter_sheet(
    file: UploadFile = File(...),
    current_staff: models.Staff = Depends(auth.require_staff),
    property_code: Optional[str] = Depends(auth.get_current_property),
    db: AsyncSession = Depends(get_db),
):
    """Upload a daily meter reading Excel per dormer (e.g. '05_DORMERS ELEC & WATER - MAY 2026').

    The workbook is expected to have one sheet per building (e.g. DT01, DT02).
    Row 2 contains headers; daily date columns are detected automatically.
    Row 1 may contain a month/year title that we also parse as a fallback.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")

    # File size validation (max 10MB)
    MAX_FILE_SIZE = 10 * 1024 * 1024
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"File too large ({len(contents) / 1024 / 1024:.1f}MB). Maximum size is 10MB.")
    if len(contents) < 100:
        raise HTTPException(status_code=400, detail="File appears to be empty or corrupted.")

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(contents), data_only=True, keep_links=False)
    except Exception as e:
        error_msg = str(e)
        if "not a zip file" in error_msg.lower() or "file format" in error_msg.lower():
            raise HTTPException(
                status_code=400,
                detail="File format not recognized. Please save as .xlsx (Excel Workbook) format. "
                       "If using Google Sheets, use File → Download → Microsoft Excel (.xlsx)."
            )
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {error_msg}")

    # Pre-load rooms and residents for matching (scoped to property if available)
    room_query = select(Room)
    if property_code:
        room_query = room_query.where(Room.property_code == property_code)
    rooms_result = await db.execute(room_query)
    rooms = rooms_result.scalars().all()
    room_by_number = {r.room_number.strip().upper(): r for r in rooms}

    # Load active residents with bed AND room info for property scoping
    resident_query = (
        select(Resident, Bed, Room)
        .join(Bed, Resident.bed_id == Bed.id, isouter=True)
        .join(Room, Bed.room_id == Room.id, isouter=True)
        .where(Resident.status == "active")
    )
    if property_code:
        resident_query = resident_query.where(
            (Room.property_code == property_code) | (Room.id.is_(None))
        )
    residents_result = await db.execute(resident_query)
    residents_rows = residents_result.all()

    resident_lookup = {}       # normalized_name -> resident
    bed_code_lookup = {}       # uppercase bed_code -> resident
    short_bed_lookup = {}      # "ROOM_NUMBER+BED_LETTER" -> resident

    for resident, bed, room in residents_rows:
        key = _normalize_name(resident.full_name)
        has_bed = bed is not None and bed.bed_code is not None
        if key and key not in resident_lookup:
            resident_lookup[key] = resident
        elif key and has_bed:
            # Prefer bed-linked resident over orphan for duplicate names
            resident_lookup[key] = resident

        if has_bed:
            # Full bed code (e.g. "DT01-401A")
            bed_code_lookup[bed.bed_code.upper()] = resident
            # Short bed code: room_number + last char of bed_code (e.g. "401A")
            if room:
                short_code = f"{room.room_number.strip().upper()}{bed.bed_code[-1].upper()}"
                short_bed_lookup[short_code] = resident

    # Build name parts index for fuzzy matching (last_name -> list of residents)
    name_parts_lookup = {}
    for norm_name, res in resident_lookup.items():
        parts = norm_name.split()
        if parts:
            last_name = parts[-1]
            if last_name not in name_parts_lookup:
                name_parts_lookup[last_name] = []
            name_parts_lookup[last_name].append(res)

    def find_resident(name: str, room_number: str = None, bed_letter: str = None):
        """Find a resident by name with bed-code fallback and fuzzy matching."""
        if not name:
            return None

        # 1. Exact normalized name match
        norm = _normalize_name(name)
        if norm in resident_lookup:
            return resident_lookup[norm]

        # 2. Bed code match (short format: room_number + bed_letter)
        if room_number and bed_letter:
            short_code = f"{room_number.strip().upper()}{bed_letter.strip().upper()}"
            if short_code in short_bed_lookup:
                return short_bed_lookup[short_code]

        # 3. Full bed code match (with property prefix)
        if property_code and room_number and bed_letter:
            prefixed_code = f"{property_code}-{room_number.strip().upper()}{bed_letter.strip().upper()}"
            if prefixed_code in bed_code_lookup:
                return bed_code_lookup[prefixed_code]

        # 4. Fuzzy: match by last name if unique
        parts = norm.split()
        if parts:
            last_name = parts[-1]
            matches = name_parts_lookup.get(last_name, [])
            if len(matches) == 1:
                return matches[0]
            # If multiple matches, try first+last name match
            if len(matches) > 1 and len(parts) >= 2:
                first_name = parts[0]
                for m in matches:
                    m_norm = _normalize_name(m.full_name)
                    m_parts = m_norm.split()
                    if m_parts and m_parts[0] == first_name:
                        return m

        return None

    total_residents_imported = 0
    total_daily_readings = 0
    all_errors = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        building = sheet_name.strip()
        if not building:
            continue

        # Auto-detect header row by finding the row with the most datetime values
        header_row_idx = None
        header_row = None
        max_dates = 0
        for r_idx in range(1, min(6, ws.max_row + 1)):
            row_vals = list(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True))[0]
            date_count = sum(1 for v in row_vals if isinstance(v, datetime))
            if date_count > max_dates:
                max_dates = date_count
                header_row_idx = r_idx
                header_row = row_vals

        if not header_row:
            all_errors.append(f"Sheet '{sheet_name}': Could not detect header row with date columns.")
            continue

        # Build header map from header row for robust column detection
        header_map = {}
        for idx, val in enumerate(header_row):
            if val is not None:
                header_map[str(val).upper().strip()] = idx

        date_cols = []  # list of (col_index, date_value)
        col_total_usage = None
        col_peso_kwh = None
        col_sub_total = None
        col_total_with_vat = None
        col_elec_bill = None
        col_water_days = None
        col_water_bill = None
        col_water_rate = None
        col_misc = {}
        header_map = {}  # map normalized header name -> column index

        for idx, val in enumerate(header_row):
            if val is None:
                continue
            if isinstance(val, datetime):
                date_cols.append((idx, val))
            elif isinstance(val, str):
                parsed_date = _parse_header_date(val)
                if parsed_date:
                    date_cols.append((idx, parsed_date))
                else:
                    vup = val.upper().strip()
                    header_map[vup] = idx
                    if vup == "TOTAL USAGE":
                        col_total_usage = idx
                    elif "# OF DAYS" in vup and "WATER" in vup:
                        col_water_days = idx
                    elif vup == "WATER BILL":
                        col_water_bill = idx
                    elif vup in {"LAUNDRY", "DRINKING WATER", "ICE CREAM", "HONESTY STORE", "COFFEE", "LOST KEYCARD", "REF RENTAL"}:
                        col_misc[vup] = idx

        if not date_cols:
            all_errors.append(f"Sheet '{sheet_name}': No date columns found in header row.")
            continue

        # Filter out stray date columns from misc sections (large gap from contiguous block)
        filtered_date_cols = []
        for idx, d in date_cols:
            if not filtered_date_cols:
                filtered_date_cols.append((idx, d))
            else:
                gap = abs((d - filtered_date_cols[-1][1]).days)
                if gap <= 60:
                    filtered_date_cols.append((idx, d))
                else:
                    break
        date_cols = filtered_date_cols

        # Determine year/month from the most common month among date columns
        from collections import Counter
        month_counts = Counter([d.month for _, d in date_cols])
        year_counts = Counter(d.year for _, d in date_cols)
        year = year_counts.most_common(1)[0][0]
        month = month_counts.most_common(1)[0][0]

        # Detect column layout based on position of 'BED' header
        # New format: room(0), flag(1), BED(2), name(3), rate(4), move_in(5), move_out(6)
        # Old format: room(0), BED(1), name(2), rate(3), move_in(4), move_out(5)
        # No BED header: same as old format (fallback)
        bed_header_idx = header_map.get("BED")
        has_flag_column = bed_header_idx is not None and bed_header_idx == 2
        if has_flag_column:
            bed_col = 2
            name_col = 3
            rate_col = 4
            move_in_col = 5
            move_out_col = 6
        else:
            bed_col = 1
            name_col = 2
            rate_col = 3
            move_in_col = 4
            move_out_col = 5

        sheet_residents = 0
        sheet_readings = 0
        current_room = None

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row:
                continue

            # Skip floor marker rows (e.g. "2nd Floor", "3rd Floor", "4F")
            if _is_floor_marker(row[0]):
                continue

            # Skip summary/total rows
            if row[0] is not None and isinstance(row[0], str):
                r0_lower = row[0].strip().lower()
                if r0_lower in ("total", "totals", "grand total", "summary", "subtotal", "sub-total"):
                    continue

            # Room number may be on its own row or repeated; forward-fill
            if row[0] is not None:
                try:
                    current_room = str(int(row[0]))
                except (ValueError, TypeError):
                    current_room = str(row[0]).strip()

            room_number = current_room
            bed_letter = str(row[bed_col]).strip() if row[bed_col] is not None else None
            resident_name = str(row[name_col]).strip() if row[name_col] is not None else None
            rate = _parse_decimal(row[rate_col]) if len(row) > rate_col else None
            move_in = _parse_date(row[move_in_col]) if len(row) > move_in_col else None
            move_out = _parse_date(row[move_out_col]) if len(row) > move_out_col else None

            if not resident_name:
                continue

            # Match resident using multi-strategy lookup
            resident = find_resident(resident_name, room_number, bed_letter)

            if not resident:
                all_errors.append(f"Sheet '{sheet_name}': Resident '{resident_name}' not found (room {room_number}, bed {bed_letter}).")
                continue

            # Extract pre-computed totals
            total_usage = _parse_decimal(row[col_total_usage]) if col_total_usage is not None and len(row) > col_total_usage else None
            peso_kwh = _parse_decimal(row[col_peso_kwh]) if col_peso_kwh is not None and len(row) > col_peso_kwh else None
            sub_total = _parse_decimal(row[col_sub_total]) if col_sub_total is not None and len(row) > col_sub_total else None
            total_with_vat = _parse_decimal(row[col_total_with_vat]) if col_total_with_vat is not None and len(row) > col_total_with_vat else None
            elec_bill = _parse_decimal(row[col_elec_bill]) if col_elec_bill is not None and len(row) > col_elec_bill else None
            water_days = None
            if col_water_days is not None and len(row) > col_water_days:
                try:
                    water_days = int(row[col_water_days])
                except (ValueError, TypeError):
                    water_days = None
            water_bill = _parse_decimal(row[col_water_bill]) if col_water_bill is not None and len(row) > col_water_bill else None

            misc_charges = {}
            for misc_name, col_idx in col_misc.items():
                if len(row) > col_idx and row[col_idx] is not None:
                    try:
                        val = Decimal(str(row[col_idx]))
                        if val > 0:
                            misc_charges[misc_name.lower().replace(" ", "_")] = str(val)
                    except Exception:
                        pass

            # Upsert summary import record
            existing_import_query = select(MeterReadingImport).where(
                MeterReadingImport.resident_id == resident.id,
                MeterReadingImport.year == year,
                MeterReadingImport.month == month,
            )
            existing_import_result = await db.execute(existing_import_query)
            existing_import = existing_import_result.scalar_one_or_none()

            water_rate_val = _parse_decimal(row[col_water_rate]) if col_water_rate is not None and len(row) > col_water_rate else None

            if existing_import:
                existing_import.building = building
                existing_import.total_electric_usage = total_usage
                existing_import.peso_kwh = peso_kwh
                existing_import.sub_total = sub_total
                existing_import.total_with_vat = total_with_vat
                existing_import.elec_bill = elec_bill
                existing_import.water_bill = water_bill
                existing_import.water_days = water_days
                existing_import.water_rate = water_rate_val
                existing_import.misc_charges = misc_charges if misc_charges else None
                existing_import.source_filename = file.filename
            else:
                imp = MeterReadingImport(
                    resident_id=resident.id,
                    building=building,
                    year=year,
                    month=month,
                    total_electric_usage=total_usage,
                    peso_kwh=peso_kwh,
                    sub_total=sub_total,
                    total_with_vat=total_with_vat,
                    elec_bill=elec_bill,
                    water_bill=water_bill,
                    water_days=water_days,
                    water_rate=water_rate_val,
                    misc_charges=misc_charges if misc_charges else None,
                    source_filename=file.filename,
                )
                db.add(imp)

            sheet_residents += 1

            # Import daily readings into meter_readings (store cumulative values for reference)
            room_id = None
            if room_number:
                room = room_by_number.get(room_number.upper())
                if room:
                    room_id = room.id

            for col_idx, col_date in date_cols:
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None:
                    continue
                try:
                    reading_val = Decimal(str(val))
                except Exception:
                    continue

                # Upsert meter reading for this date
                existing_reading_query = select(MeterReading).where(
                    MeterReading.resident_id == resident.id,
                    MeterReading.reading_date == col_date.date() if isinstance(col_date, datetime) else col_date,
                )
                existing_reading_result = await db.execute(existing_reading_query)
                existing_reading = existing_reading_result.scalar_one_or_none()

                if existing_reading:
                    existing_reading.electric_reading = reading_val
                    existing_reading.building = building
                    existing_reading.room_id = room_id
                else:
                    reading = MeterReading(
                        building=building,
                        room_id=room_id,
                        resident_id=resident.id,
                        reading_date=col_date.date() if isinstance(col_date, datetime) else col_date,
                        electric_reading=reading_val,
                        water_reading=None,
                        status="estimated",
                        variance_pct=Decimal("0"),
                    )
                    db.add(reading)
                sheet_readings += 1

        total_residents_imported += sheet_residents
        total_daily_readings += sheet_readings

    await db.commit()
    return MeterReadingDailySheetResult(
        building=", ".join(wb.sheetnames),
        year=year,
        month=month,
        residents_imported=total_residents_imported,
        daily_readings_imported=total_daily_readings,
        errors=all_errors,
        message=f"Imported {total_residents_imported} residents with {total_daily_readings} daily readings across {len(wb.sheetnames)} sheet(s)."
    )


@router.get("/meter-readings", response_model=List[MeterReadingOut])
async def list_meter_readings(
    building: Optional[str] = Query(None),
    resident_id: Optional[UUID] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    limit: int = Query(1000, ge=1, le=5000),
    current_staff: models.Staff = Depends(auth.require_staff),
    property_code: Optional[str] = Depends(auth.get_current_property),
    db: AsyncSession = Depends(get_db),
):
    query = (
        select(MeterReading, Room, Resident, Bed)
        .join(Room, MeterReading.room_id == Room.id, isouter=True)
        .join(Resident, MeterReading.resident_id == Resident.id, isouter=True)
        .join(Bed, Resident.bed_id == Bed.id, isouter=True)
        .order_by(MeterReading.reading_date.desc())
    )
    if property_code:
        query = query.where(Room.property_code == property_code)
    if building:
        query = query.where(MeterReading.building == building)
    if resident_id:
        query = query.where(MeterReading.resident_id == resident_id)
    if year and month:
        start = date(year, month, 1)
        if month == 12:
            end = date(year + 1, 1, 1)
        else:
            end = date(year, month + 1, 1)
        query = query.where(MeterReading.reading_date >= start, MeterReading.reading_date < end)

    query = query.limit(limit)
    result = await db.execute(query)
    rows = result.all()

    out = []
    for reading, room, resident, bed in rows:
        out.append(MeterReadingOut(
            id=reading.id,
            building=reading.building,
            room_id=reading.room_id,
            resident_id=reading.resident_id,
            reading_date=reading.reading_date,
            electric_reading=reading.electric_reading,
            water_reading=reading.water_reading,
            status=reading.status,
            variance_pct=reading.variance_pct,
            room_number=room.room_number if room else None,
            resident_name=resident.full_name if resident else None,
            bed_code=bed.bed_code if bed else None,
        ))
    return out


@router.get("/meter-readings/daily-grid", response_model=MeterReadingDailyGridOut)
async def get_daily_meter_grid(
    year: int = Query(...),
    month: int = Query(...),
    building: Optional[str] = Query(None),
    current_staff: models.Staff = Depends(auth.require_staff),
    property_code: Optional[str] = Depends(auth.get_current_property),
    db: AsyncSession = Depends(get_db),
):
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    days_in_month = calendar.monthrange(year, month)[1]

    # Fetch active residents with bed/room info
    query = (
        select(Resident, Bed, Room)
        .join(Bed, Resident.bed_id == Bed.id, isouter=True)
        .join(Room, Bed.room_id == Room.id, isouter=True)
        .where(Resident.status == "active")
    )
    if property_code:
        query = query.where(Room.property_code == property_code)
    if building:
        query = query.where(Room.building == building)
    query = query.order_by(Room.room_number.asc().nullslast(), Bed.bed_number.asc().nullslast())
    result = await db.execute(query)
    rows = result.all()

    # Fetch meter readings for the period
    reading_query = select(MeterReading).where(
        MeterReading.reading_date >= start_date,
        MeterReading.reading_date < end_date,
        MeterReading.resident_id.isnot(None),
    )
    if property_code:
        reading_query = reading_query.join(Room, MeterReading.room_id == Room.id, isouter=True).where(Room.property_code == property_code)
    if building:
        reading_query = reading_query.where(MeterReading.building == building)
    reading_result = await db.execute(reading_query)
    readings = reading_result.scalars().all()

    # Fetch MeterReadingImport records for this period
    import_query = select(MeterReadingImport).where(
        MeterReadingImport.year == year,
        MeterReadingImport.month == month,
    )
    if property_code:
        _prop_buildings = await get_property_buildings(db, property_code)
        if _prop_buildings:
            import_query = import_query.where(MeterReadingImport.building.in_(_prop_buildings))
    if building:
        import_query = import_query.where(MeterReadingImport.building == building)
    import_result = await db.execute(import_query)
    imports = import_result.scalars().all()

    # Index imports by resident_id
    imports_map = {}
    latest_import = None
    for imp in imports:
        imports_map[str(imp.resident_id)] = imp
        if latest_import is None or imp.created_at > latest_import.created_at:
            latest_import = imp

    # Index readings by resident_id and date
    readings_map = {}
    for r in readings:
        rid = str(r.resident_id)
        dkey = r.reading_date.strftime("%Y-%m-%d")
        if rid not in readings_map:
            readings_map[rid] = {}
        readings_map[rid][dkey] = MeterReadingDailyCell(
            reading_id=r.id,
            electric_reading=r.electric_reading,
            water_reading=r.water_reading,
            status=r.status,
        )

    # Track which beds have active residents
    occupied_bed_ids = set()

    residents_out = []
    for resident, bed, room in rows:
        rid = str(resident.id)
        days = _count_days_stayed(resident.move_in_date, resident.move_out_date, start_date, end_date)
        bed_letter = None
        if bed and bed.bed_code:
            bed_letter = bed.bed_code[-1] if len(bed.bed_code) > 0 else None
            occupied_bed_ids.add(str(bed.id))

        # Merge import data if available
        imp = imports_map.get(rid)

        residents_out.append(MeterReadingDailyRow(
            resident_id=resident.id,
            resident_name=resident.full_name,
            room_number=room.room_number if room else None,
            bed_code=bed.bed_code if bed else None,
            bed_letter=bed_letter,
            monthly_rate=resident.monthly_rate,
            move_in_date=resident.move_in_date,
            move_out_date=resident.move_out_date,
            days_in_month=days,
            readings=readings_map.get(rid, {}),
            # Import summary data
            total_electric_usage=imp.total_electric_usage if imp else None,
            peso_kwh=imp.peso_kwh if imp else None,
            sub_total=imp.sub_total if imp else None,
            total_with_vat=imp.total_with_vat if imp else None,
            elec_bill=imp.elec_bill if imp else None,
            water_bill=imp.water_bill if imp else None,
            water_days=imp.water_days if imp else None,
            water_rate=imp.water_rate if imp else None,
            misc_charges=imp.misc_charges if imp else None,
            source_filename=imp.source_filename if imp else None,
        ))

    # Find vacant beds (beds with no active resident assigned)
    vacant_beds = []
    all_beds_query = (
        select(Bed, Room)
        .join(Room, Bed.room_id == Room.id)
    )
    if property_code:
        all_beds_query = all_beds_query.where(Room.property_code == property_code)
    if building:
        all_beds_query = all_beds_query.where(Room.building == building)
    all_beds_query = all_beds_query.order_by(Room.room_number.asc(), Bed.bed_number.asc())
    all_beds_result = await db.execute(all_beds_query)
    for bed_item, room_item in all_beds_result.all():
        if str(bed_item.id) not in occupied_bed_ids:
            bed_letter_v = None
            if bed_item.bed_code:
                bed_letter_v = bed_item.bed_code[-1] if len(bed_item.bed_code) > 0 else None
            vacant_beds.append(VacantBedRow(
                room_number=room_item.room_number if room_item else "",
                bed_code=bed_item.bed_code,
                bed_letter=bed_letter_v,
                bed_number=bed_item.bed_number,
            ))

    # Build import info
    import_info = None
    if latest_import:
        import_info = ImportInfo(
            source_filename=latest_import.source_filename,
            imported_at=latest_import.created_at,
            resident_count=len(imports),
        )

    return MeterReadingDailyGridOut(
        year=year,
        month=month,
        days_in_month=days_in_month,
        residents=residents_out,
        vacant_beds=vacant_beds,
        import_info=import_info,
        water_config={},
    )


@router.post("/meter-readings/bulk-upsert")
async def bulk_upsert_meter_readings(
    data: List[MeterReadingCreate],
    current_staff: models.Staff = Depends(auth.require_staff),
    db: AsyncSession = Depends(get_db),
):
    """Bulk upsert daily meter readings (used by the grid UI)."""
    created = 0
    updated = 0
    for item in data:
        if not item.resident_id or not item.reading_date:
            continue

        # Check if reading already exists for this resident + date
        existing_query = select(MeterReading).where(
            MeterReading.resident_id == item.resident_id,
            MeterReading.reading_date == item.reading_date,
        )
        existing_result = await db.execute(existing_query)
        existing = existing_result.scalar_one_or_none()

        if existing:
            if item.electric_reading is not None:
                existing.electric_reading = item.electric_reading
            if item.water_reading is not None:
                existing.water_reading = item.water_reading
            updated += 1
        else:
            reading = MeterReading(
                building=item.building,
                room_id=item.room_id,
                resident_id=item.resident_id,
                reading_date=item.reading_date,
                electric_reading=item.electric_reading,
                water_reading=item.water_reading,
                variance_pct=Decimal("0"),
                status="approved",
            )
            db.add(reading)
            created += 1

    await db.commit()
    return {"created": created, "updated": updated}


@router.post("/preview", response_model=BillingPreviewResponse)
async def preview_billings(
    data: BillingGenerateRequest,
    current_staff: models.Staff = Depends(auth.require_staff),
    property_code: Optional[str] = Depends(auth.get_current_property),
    db: AsyncSession = Depends(get_db),
):
    residents_by_room, no_room, all_rows = await _get_active_residents_with_rooms(db, data.building, property_code)
    if not all_rows:
        detail = f"No active residents found" + (f" for building '{data.building}'" if data.building else "")
        raise HTTPException(status_code=400, detail=detail)

    resident_electric = await _compute_resident_electric(db, data.billing_period, data.building, property_code)
    resident_water, total_days, rate_per_day = await _compute_water_by_days(
        db, data.billing_period, data.total_water_bill, data.building, property_code
    )
    resident_other, total_other = await _compute_other_charges(
        db, data.billing_period, data.other_charges, data.building, property_code
    )
    previous_balances = await _compute_previous_balances(db, data.billing_period, data.building, property_code)

    total_residents = len(all_rows)

    preview_rows = []
    total_rent = Decimal("0")
    total_electric = Decimal("0")
    total_water = Decimal("0")
    total_other_computed = Decimal("0")
    total_all = Decimal("0")

    # Residents with rooms
    for room_id, room_data in residents_by_room.items():
        room = room_data["room"]
        room_residents = room_data["residents"]

        for resident, bed in room_residents:
            rid = str(resident.id)
            elec = resident_electric.get(rid, Decimal("0"))
            wat = resident_water.get(rid, Decimal("0"))
            other = resident_other.get(rid, Decimal("0"))
            prev_bal = previous_balances.get(rid, Decimal("0"))
            total = resident.monthly_rate + elec + wat + other + prev_bal
            preview_rows.append(BillingPreviewRow(
                resident_id=resident.id,
                resident_name=resident.full_name,
                room_number=room.room_number if room else None,
                bed_number=bed.bed_number if bed else None,
                bed_code=bed.bed_code if bed else None,
                monthly_rate=resident.monthly_rate,
                rent_amount=resident.monthly_rate,
                electric_charge=elec,
                water_charge=wat,
                other_charges=other,
                previous_balance=prev_bal,
                total_amount=total,
            ))
            total_rent += resident.monthly_rate
            total_electric += elec
            total_water += wat
            total_other_computed += other
            total_all += total

    # Residents without a room
    for resident, bed in no_room:
        rid = str(resident.id)
        elec = resident_electric.get(rid, Decimal("0"))
        wat = resident_water.get(rid, Decimal("0"))
        other = resident_other.get(rid, Decimal("0"))
        prev_bal = previous_balances.get(rid, Decimal("0"))
        total = resident.monthly_rate + elec + wat + other + prev_bal
        preview_rows.append(BillingPreviewRow(
            resident_id=resident.id,
            resident_name=resident.full_name,
            room_number=None,
            bed_number=bed.bed_number if bed else None,
            bed_code=bed.bed_code if bed else None,
            monthly_rate=resident.monthly_rate,
            rent_amount=resident.monthly_rate,
            electric_charge=elec,
            water_charge=wat,
            other_charges=other,
            previous_balance=prev_bal,
            total_amount=total,
        ))
        total_rent += resident.monthly_rate
        total_electric += elec
        total_water += wat
        total_other_computed += other
        total_all += total

    other_per_head = total_other_computed / total_residents if total_residents > 0 else Decimal("0")

    return BillingPreviewResponse(
        billing_period=data.billing_period,
        building=data.building,
        total_residents=total_residents,
        rows=preview_rows,
        summary={
            "total_rent": str(total_rent),
            "total_electric": str(total_electric),
            "total_water": str(total_water),
            "total_other": str(total_other_computed),
            "grand_total": str(total_all),
            "electric_per_head": str(total_electric / total_residents if total_residents > 0 else Decimal("0")),
            "water_per_head": str(total_water / total_residents if total_residents > 0 else Decimal("0")),
            "other_per_head": str(other_per_head),
            "total_days": total_days,
            "water_rate_per_day": str(rate_per_day),
        }
    )


@router.get("/import-status", response_model=BillingImportStatusOut)
async def get_billing_import_status(
    billing_period: str = Query(..., description="YYYY-MM"),
    building: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    year, month = map(int, billing_period.split("-"))
    query = select(MeterReadingImport).where(
        MeterReadingImport.year == year,
        MeterReadingImport.month == month,
    )
    if building:
        query = query.where(MeterReadingImport.building == building)

    result = await db.execute(query)
    imports = result.scalars().all()

    if not imports:
        return BillingImportStatusOut(
            billing_period=billing_period,
            building=building,
            has_imports=False,
            import_count=0,
        )

    total_water = Decimal("0")
    total_misc = Decimal("0")
    total_electric = Decimal("0")
    source_filename = None

    for imp in imports:
        if imp.water_bill:
            total_water += imp.water_bill
        if imp.total_electric_usage:
            total_electric += imp.total_electric_usage
        if imp.misc_charges:
            for key, val in imp.misc_charges.items():
                try:
                    total_misc += Decimal(str(val))
                except Exception:
                    pass
        if imp.source_filename:
            source_filename = imp.source_filename

    return BillingImportStatusOut(
        billing_period=billing_period,
        building=building,
        has_imports=True,
        import_count=len(imports),
        total_imported_water=total_water,
        total_imported_misc=total_misc,
        total_imported_electric=total_electric,
        source_filename=source_filename,
    )


@router.get("/", response_model=List[BillingWithResidentOut])
async def list_billings(
    resident_id: Optional[UUID] = Query(None),
    status: Optional[str] = Query(None),
    current_staff: models.Staff = Depends(auth.require_staff),
    property_code: Optional[str] = Depends(auth.get_current_property),
    db: AsyncSession = Depends(get_db),
):
    query = (
        select(Billing, Resident, Bed, Room)
        .join(Resident, Billing.resident_id == Resident.id)
        .join(Bed, Resident.bed_id == Bed.id, isouter=True)
        .join(Room, Bed.room_id == Room.id, isouter=True)
    )
    if property_code:
        query = query.where(Room.property_code == property_code)
    if resident_id:
        query = query.where(Billing.resident_id == resident_id)
    if status:
        query = query.where(Billing.status == status)
    query = query.order_by(Billing.created_at.desc())
    result = await db.execute(query)
    rows = result.all()

    out = []
    for billing, resident, bed, room in rows:
        out.append(BillingWithResidentOut(
            id=billing.id,
            resident_id=billing.resident_id,
            billing_period=billing.billing_period,
            rent_amount=billing.rent_amount,
            electric_charge=billing.electric_charge,
            water_charge=billing.water_charge,
            other_charges=billing.other_charges,
            total_amount=billing.total_amount,
            status=billing.status,
            created_at=billing.created_at,
            resident_name=resident.full_name if resident else None,
            bed_code=bed.bed_code if bed else None,
            room_number=room.room_number if room else None,
        ))
    return out
